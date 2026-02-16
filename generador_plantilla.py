"""
===========================================
Automatizador de Nomenclatura de Archivos
===========================================

Este script automatiza la creación y nomenclatura de archivos Excel para el ejercicio de coberturas,
implementando un sistema estandarizado basado en códigos de países y categorías.

Funcionalidades Principales:
--------------------------
1. Selección asistida de país y obtención de su código
2. Búsqueda y selección de categorías de productos
3. Generación automática de nombres de archivo estandarizados
4. Creación de archivos Excel con estructura predefinida

Componentes Principales:
----------------------
- Diccionario de países (countries): Mapeo de países y sus códigos
- Lista de categorías (categorias): Categorías de productos con sus códigos
- Clase Colors: Códigos ANSI para formato de texto en consola

Funciones Principales:
--------------------
- obtener_codigo_pais(): Obtiene el código del país desde la entrada del usuario
- buscar_categorias(): Busca categorías basadas en palabras clave
- seleccionar_categoria(): Maneja la selección de categoría por el usuario
- crear_excel(): Genera el archivo Excel con la estructura requerida
- main(): Función principal que coordina el flujo del programa

Ejemplo de Uso:
-------------
1. Ejecutar el script
2. Ingresar país (ej: "mexico" o "mex")
3. Buscar categoría por palabra clave
4. Seleccionar categoría de la lista
5. Ingresar nombre del fabricante
6. El script generará automáticamente el archivo Excel con el nombre estandarizado

Formato del nombre de archivo resultante:
[código_país]_[código_categoría]_[fabricante].xlsx
Ejemplo: 52_SODA_COCACOLA.xlsx
"""
# script_optimizado.py
import os
import re
import unicodedata
import sys
try:
    from rich.console import Console
except ImportError as exc:
    raise ImportError("Falta la dependencia 'rich'. Instala con: pip install rich") from exc

console = Console(highlight=False, soft_wrap=True)

def print(*args, sep=' ', end='\n', **kwargs):
    """Wrapper de salida usando rich para una lectura mas clara en terminal."""
    text = sep.join(str(arg) for arg in args)
    style = kwargs.pop('style', None)
    markup = kwargs.pop('markup', True)
    console.print(text, style=style, end=end, markup=markup)

def input(prompt=''):
    """Wrapper de entrada usando rich para prompts coloreados."""
    return console.input(str(prompt))

# Estilos de texto (markup rich).
class Colors:
    HEADER = '[bold bright_cyan]'
    OKBLUE = '[bold #4FC3F7]'
    OKCYAN = '[bold #80DEEA]'
    OKGREEN = '[bold #66BB6A]'
    WARNING = '[bold yellow]'
    FAIL = '[bold red]'
    ENDC = '[/]'
    BOLD = ''
    UNDERLINE = ''

class PromptColors:
    COUNTRY = '[bold #EC407A]'
    CATEGORY_SEARCH = '[bold #42A5F5]'
    CATEGORY_SELECT = '[bold #26C6DA]'
    MANUFACTURER = '[bold #FFA726]'
    BRAND = '[bold #AB47BC]'
    TEMPLATE_SELECT = '[bold #5C6BC0]'
    CATEGORY_LABEL = '[bold #26A69A]'
    PLAYERS_LABEL = '[bold #90A4AE]'
    DISTRIBUTION_LABEL = '[bold #EF5350]'
    TREE_UNIT = '[bold #8D6E63]'
    CONTINUE = '[bold #78909C]'

TREE_UNIT_MAP = {
    "U": "Units",
    "L": "Litros",
    "K": "Kilos",
    "T": "Toneladas",
    "R": "Rollos",
    "M": "Metros",
    "H": "Hojas",
}

def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s)
                   if unicodedata.category(c) != 'Mn').lower()

countries = {
    'latam': '10', 'lat': '10',
    'argentina': '54', 'arg': '54',
    'bolivia': '91', 'bol': '91',
    'brasil': '55', 'bra': '55', 'brazil': '55',
    'cam': '12',
    'chile': '56', 'chl': '56',
    'colombia': '57', 'col': '57',
    'ecuador': '93', 'ecu': '93',
    'guatemala': '62', 'gt': '62', 'gtm': '62',
    'el salvador': '63', 'elsalvador': '63', 'slv': '63',
    'honduras': '64', 'hnd': '64',
    'nicaragua': '65', 'nic': '65',
    'costa rica': '66', 'costarica': '66', 'cri': '66',
    'panama': '67', 'pan': '67',
    'republica dominicana': '69', 'republicadominicana': '69', 'repdom': '69', 'rd': '69', 'dom': '69',
    'mexico': '52', 'mex': '52', 'mx': '52',
    'peru': '51', 'per': '51'
}

pop_coverage = {
    'Argentina': '90%',
    'Bolivia': '60%',
    'Brasil': '82%',
    'Chile': '78%',
    'Colombia': '65%',
    'Ecuador': '55%',
    'Mexico': '64%',
    'Peru': '66%',
    'CAM': '74%',
    'Costa Rica': '94%',
    'El Salvador': '85%',
    'Guatemala': '69%',
    'Honduras': '65%',
    'Nicaragua': '57%',
    'Panama': '92%',
    'Republica Dominicana': '63.29%'
}

def obtener_nombre_pais(key):
    mapping = {
        'latam': 'LatAm', 'lat': 'LatAm',
        'argentina': 'Argentina', 'arg': 'Argentina',
        'bolivia': 'Bolivia', 'bol': 'Bolivia',
        'brasil': 'Brasil', 'bra': 'Brasil', 'brazil': 'Brasil',
        'cam': 'CAM',
        'chile': 'Chile', 'chl': 'Chile',
        'colombia': 'Colombia', 'col': 'Colombia',
        'ecuador': 'Ecuador', 'ecu': 'Ecuador',
        'guatemala': 'Guatemala', 'gt': 'Guatemala', 'gtm': 'Guatemala',
        'el salvador': 'El Salvador', 'elsalvador': 'El Salvador', 'slv': 'El Salvador',
        'honduras': 'Honduras', 'hnd': 'Honduras',
        'nicaragua': 'Nicaragua', 'nic': 'Nicaragua',
        'costa rica': 'Costa Rica', 'costarica': 'Costa Rica', 'cri': 'Costa Rica',
        'panama': 'Panama', 'pan': 'Panama',
        'republica dominicana': 'Republica Dominicana', 'republicadominicana': 'Republica Dominicana',
        'repdom': 'Republica Dominicana', 'rd': 'Republica Dominicana', 'dom': 'Republica Dominicana',
        'mexico': 'Mexico', 'mex': 'Mexico', 'mx': 'Mexico',
        'peru': 'Peru', 'per': 'Peru'
    }
    return mapping.get(key, "Desconocido")

def obtener_codigo_pais(input_pais):
    input_normalizado = strip_accents(input_pais)
    if input_normalizado in countries:
        return countries[input_normalizado], obtener_nombre_pais(input_normalizado)
    else:
        # Intentar buscar por substring
        matches = [key for key in countries.keys() if input_normalizado in key]
        if len(matches) == 1:
            return countries[matches[0]], obtener_nombre_pais(matches[0])
        elif len(matches) > 1:
            print(f"{Colors.WARNING}Se encontraron múltiples países:{Colors.ENDC}")
            for idx, key in enumerate(matches, 1):
                print(f"{Colors.OKBLUE}{idx}. {obtener_nombre_pais(key)}{Colors.ENDC}")
            try:
                seleccion = int(input(f"{PromptColors.COUNTRY}Seleccione el país (número): {Colors.ENDC}"))
                if 1 <= seleccion <= len(matches):
                    selected_key = matches[seleccion - 1]
                    return countries[selected_key], obtener_nombre_pais(selected_key)
                else:
                    print(f"{Colors.FAIL}Selección inválida.{Colors.ENDC}")
                    return None, None
            except ValueError:
                print(f"{Colors.FAIL}Entrada inválida.{Colors.ENDC}")
                return None, None
        else:
            return None, None

# Lista de categorías con sus descripciones y códigos
categorias = [
    {'categoria': 'Alimentos', 'descripcion': 'Carne Fresca', 'cod': 'MEAT'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Pañitos + Pañales', 'cod': 'CRDT'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Snacks)', 'cod': 'CRSN'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Alcohólicas', 'cod': 'ALCB'},
    {'categoria': 'Bebidas', 'descripcion': 'Cervezas', 'cod': 'BEER'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Gaseosas', 'cod': 'CARB'},
    {'categoria': 'Bebidas', 'descripcion': 'Agua Gasificada', 'cod': 'CWAT'},
    {'categoria': 'Bebidas', 'descripcion': 'Água de Coco', 'cod': 'COCW'},
    {'categoria': 'Bebidas', 'descripcion': 'Café_Consolidado de Café', 'cod': 'COFF'},
    {'categoria': 'Bebidas', 'descripcion': 'Cross Category (Bebidas)', 'cod': 'CRBE'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Energéticas', 'cod': 'ENDR'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Saborizadas Sin Gas', 'cod': 'FLBE'},
    {'categoria': 'Bebidas', 'descripcion': 'Café Tostado y Molido', 'cod': 'GCOF'},
    {'categoria': 'Bebidas', 'descripcion': 'Jugos Caseros', 'cod': 'HJUI'},
    {'categoria': 'Bebidas', 'descripcion': 'Té Helado', 'cod': 'ITEA'},
    {'categoria': 'Bebidas', 'descripcion': 'Café Instantáneo_Café Sucedáneo', 'cod': 'ICOF'},
    {'categoria': 'Bebidas', 'descripcion': 'Jugos y Nectares', 'cod': 'JUNE'},
    {'categoria': 'Bebidas', 'descripcion': 'Zumos de Vegetales', 'cod': 'VEJU'},
    {'categoria': 'Bebidas', 'descripcion': 'Agua Natural', 'cod': 'WATE'},
    {'categoria': 'Bebidas', 'descripcion': 'Gaseosas + Aguas', 'cod': 'CSDW'},
    {'categoria': 'Bebidas', 'descripcion': 'Mixta Café+Malta', 'cod': 'MXCM'},
    {'categoria': 'Bebidas', 'descripcion': 'Mixta Dolce Gusto_Mixta Té Helado + Café + Modificadores', 'cod': 'MXDG'},
    {'categoria': 'Bebidas', 'descripcion': 'Mixta Jugos y Leches', 'cod': 'MXJM'},
    {'categoria': 'Bebidas', 'descripcion': 'Mixta Jugos Líquidos + Bebidas de Soja', 'cod': 'MXJS'},
    {'categoria': 'Bebidas', 'descripcion': 'Mixta Té+Café', 'cod': 'MXTC'},
    {'categoria': 'Bebidas', 'descripcion': 'Jugos Liquidos_Jugos Polvo', 'cod': 'JUIC'},
    {'categoria': 'Bebidas', 'descripcion': 'Refrescos en Polvo_Jugos _ Bebidas Instantáneas En Polvo _ Jugos Polvo', 'cod': 'PWDJ'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Refrescantes', 'cod': 'RFDR'},
    {'categoria': 'Bebidas', 'descripcion': 'Refrescos Líquidos_Jugos Líquidos', 'cod': 'RTDJ'},
    {'categoria': 'Bebidas', 'descripcion': 'Té Líquido _ Listo para Tomar', 'cod': 'RTEA'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas de Soja', 'cod': 'SOYB'},
    {'categoria': 'Bebidas', 'descripcion': 'Bebidas Isotónicas', 'cod': 'SPDR'},
    {'categoria': 'Bebidas', 'descripcion': 'Té e Infusiones_Te_Infusión Hierbas', 'cod': 'TEAA'},
    {'categoria': 'Bebidas', 'descripcion': 'Yerba Mate', 'cod': 'YERB'},
    {'categoria': 'Lacteos', 'descripcion': 'Manteca', 'cod': 'BUTT'},
    {'categoria': 'Lacteos', 'descripcion': 'Queso Fresco y para Untar', 'cod': 'CHEE'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Condensada', 'cod': 'CMLK'},
    {'categoria': 'Lacteos', 'descripcion': 'Queso Untable', 'cod': 'CRCH'},
    {'categoria': 'Lacteos', 'descripcion': 'Yoghurt p_beber', 'cod': 'DYOG'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Culinaria_Leche Evaporada', 'cod': 'EMLK'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Fermentada', 'cod': 'FRMM'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Líquida Saborizada_Leche Líquida Con Sabor', 'cod': 'FMLK'},
    {'categoria': 'Lacteos', 'descripcion': 'Fórmulas Infantiles', 'cod': 'FRMK'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Líquida', 'cod': 'LQDM'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche Larga Vida', 'cod': 'LLFM'},
    {'categoria': 'Lacteos', 'descripcion': 'Margarina', 'cod': 'MARG'},
    {'categoria': 'Lacteos', 'descripcion': 'Queso Fundido', 'cod': 'MCHE'},
    {'categoria': 'Lacteos', 'descripcion': 'Crema de Leche', 'cod': 'MKCR'},
    {'categoria': 'Lacteos', 'descripcion': 'Mixta Lácteos_Postre+Leches+Yogurt', 'cod': 'MXDI'},
    {'categoria': 'Lacteos', 'descripcion': 'Mixta Leches', 'cod': 'MXMI'},
    {'categoria': 'Lacteos', 'descripcion': 'Mixta Yoghurt+Postres', 'cod': 'MXYD'},
    {'categoria': 'Lacteos', 'descripcion': 'Petit Suisse', 'cod': 'PTSS'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche en Polvo', 'cod': 'PWDM'},
    {'categoria': 'Lacteos', 'descripcion': 'Yoghurt p_comer', 'cod': 'SYOG'},
    {'categoria': 'Lacteos', 'descripcion': 'Leche_Leche Líquida Blanca _ Leche Liq. Natural', 'cod': 'MILK'},
    {'categoria': 'Lacteos', 'descripcion': 'Yoghurt', 'cod': 'YOGH'},
    {'categoria': 'Ropas y Calzados', 'descripcion': 'Ropas', 'cod': 'CLOT'},
    {'categoria': 'Ropas y Calzados', 'descripcion': 'Calzados', 'cod': 'FOOT'},
    {'categoria': 'Ropas y Calzados', 'descripcion': 'Medias_Calcetines', 'cod': 'SOCK'},
    {'categoria': 'Alimentos', 'descripcion': 'Arepas', 'cod': 'AREP'},
    {'categoria': 'Alimentos', 'descripcion': 'Cereales Infantiles', 'cod': 'BCER'},
    {'categoria': 'Alimentos', 'descripcion': 'Nutrición Infantil_Colados y Picados', 'cod': 'BABF'},
    {'categoria': 'Alimentos', 'descripcion': 'Frijoles', 'cod': 'BEAN'},
    {'categoria': 'Alimentos', 'descripcion': 'Galletas', 'cod': 'BISC'},
    {'categoria': 'Alimentos', 'descripcion': 'Caldos_Caldos y Sazonadores', 'cod': 'BOUI'},
    {'categoria': 'Alimentos', 'descripcion': 'Pan', 'cod': 'BREA'},
    {'categoria': 'Alimentos', 'descripcion': 'Apanados_Empanizadores', 'cod': 'BRCR'},
    {'categoria': 'Alimentos', 'descripcion': 'Empanados', 'cod': 'BRDC'},
    {'categoria': 'Alimentos', 'descripcion': 'Cereales_Cereales Desayuno_Avenas y Cereales', 'cod': 'CERE'},
    {'categoria': 'Alimentos', 'descripcion': 'Hamburguesas', 'cod': 'BURG'},
    {'categoria': 'Alimentos', 'descripcion': 'Mezclas Listas para Tortas_Preparados Base Harina Trigo', 'cod': 'CCMX'},
    {'categoria': 'Alimentos', 'descripcion': 'Queques_Ponques Industrializados', 'cod': 'CAKE'},
    {'categoria': 'Alimentos', 'descripcion': 'Conservas De Pescado', 'cod': 'FISH'},
    {'categoria': 'Alimentos', 'descripcion': 'Conservas de Frutas y Verduras', 'cod': 'CFAV'},
    {'categoria': 'Alimentos', 'descripcion': 'Dulce de Leche_Manjar', 'cod': 'CRML'},
    {'categoria': 'Alimentos', 'descripcion': 'Alfajores', 'cod': 'CMLC'},
    {'categoria': 'Alimentos', 'descripcion': 'Barras de Cereal', 'cod': 'CBAR'},
    {'categoria': 'Alimentos', 'descripcion': 'Pollo', 'cod': 'CHCK'},
    {'categoria': 'Alimentos', 'descripcion': 'Chocolate', 'cod': 'CHOC'},
    {'categoria': 'Alimentos', 'descripcion': 'Chocolate de Taza_Achocolatados _ Cocoas', 'cod': 'COCO'},
    {'categoria': 'Alimentos', 'descripcion': 'Salsas Frías', 'cod': 'COLS'},
    {'categoria': 'Alimentos', 'descripcion': 'Compotas', 'cod': 'COMP'},
    {'categoria': 'Alimentos', 'descripcion': 'Condimentos y Especias', 'cod': 'SPIC'},
    {'categoria': 'Alimentos', 'descripcion': 'Chocolate de Mesa', 'cod': 'CKCH'},
    {'categoria': 'Alimentos', 'descripcion': 'Aceite_Aceites Comestibles', 'cod': 'COIL'},
    {'categoria': 'Alimentos', 'descripcion': 'Salsas Listas_Salsas Caseras Envasadas', 'cod': 'CSAU'},
    {'categoria': 'Alimentos', 'descripcion': 'Grano, Harina y Masa de Maíz', 'cod': 'CNML'},
    {'categoria': 'Alimentos', 'descripcion': 'Fécula de Maíz', 'cod': 'CNST'},
    {'categoria': 'Alimentos', 'descripcion': 'Harina De Maíz', 'cod': 'CNFL'},
    {'categoria': 'Alimentos', 'descripcion': 'Ayudantes Culinarios', 'cod': 'CAID'},
    {'categoria': 'Alimentos', 'descripcion': 'Postres Preparados', 'cod': 'DESS'},
    {'categoria': 'Alimentos', 'descripcion': 'Jamón Endiablado', 'cod': 'DHAM'},
    {'categoria': 'Alimentos', 'descripcion': 'Semillas y Frutos Secos', 'cod': 'DFNS'},
    {'categoria': 'Alimentos', 'descripcion': 'Pan de Pascua', 'cod': 'EBRE'},
    {'categoria': 'Alimentos', 'descripcion': 'Huevos de Páscua', 'cod': 'EEGG'},
    {'categoria': 'Alimentos', 'descripcion': 'Huevos', 'cod': 'EGGS'},
    {'categoria': 'Alimentos', 'descripcion': 'Flash Cecinas', 'cod': 'FLSS'},
    {'categoria': 'Alimentos', 'descripcion': 'Harinas', 'cod': 'FLOU'},
    {'categoria': 'Alimentos', 'descripcion': 'Platos Listos Congelados', 'cod': 'FRDS'},
    {'categoria': 'Alimentos', 'descripcion': 'Alimentos Congelados', 'cod': 'FRFO'},
    {'categoria': 'Alimentos', 'descripcion': 'Jamones', 'cod': 'HAMS'},
    {'categoria': 'Alimentos', 'descripcion': 'Cereales Calientes_Cereales Precocidos', 'cod': 'HCER'},
    {'categoria': 'Alimentos', 'descripcion': 'Salsas Picantes', 'cod': 'HOTS'},
    {'categoria': 'Alimentos', 'descripcion': 'Helados', 'cod': 'ICEC'},
    {'categoria': 'Alimentos', 'descripcion': 'Pan Industrializado', 'cod': 'IBRE'},
    {'categoria': 'Alimentos', 'descripcion': 'Puré Instantáneo', 'cod': 'IMPO'},
    {'categoria': 'Alimentos', 'descripcion': 'Fideos Instantáneos', 'cod': 'INOO'},
    {'categoria': 'Alimentos', 'descripcion': 'Mermeladas', 'cod': 'JAMS'},
    {'categoria': 'Alimentos', 'descripcion': 'Ketchup', 'cod': 'KETC'},
    {'categoria': 'Alimentos', 'descripcion': 'Jugo de Limon Adereso', 'cod': 'LJDR'},
    {'categoria': 'Alimentos', 'descripcion': 'Maltas', 'cod': 'MALT'},
    {'categoria': 'Alimentos', 'descripcion': 'Adobos _ Sazonadores', 'cod': 'SEAS'},
    {'categoria': 'Alimentos', 'descripcion': 'Mayonesa', 'cod': 'MAYO'},
    {'categoria': 'Alimentos', 'descripcion': 'Cárnicos', 'cod': 'MEAT'},
    {'categoria': 'Alimentos', 'descripcion': 'Modificadores de Leche_Saborizadores p_leche', 'cod': 'MLKM'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Cereales Infantiles+Avenas', 'cod': 'MXCO'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Caldos + Saborizantes', 'cod': 'MXBS'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Caldos + Sopas', 'cod': 'MXSB'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Cereales + Cereales Calientes', 'cod': 'MXCH'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Chocolate + Manjar', 'cod': 'MXCC'},
    {'categoria': 'Alimentos', 'descripcion': 'Galletas, snacks y mini tostadas', 'cod': 'MXSN'},
    {'categoria': 'Alimentos', 'descripcion': 'Aceites + Mantecas', 'cod': 'COBT'},
    {'categoria': 'Alimentos', 'descripcion': 'Aceites + Conservas De Pescado', 'cod': 'COCF'},
    {'categoria': 'Alimentos', 'descripcion': 'Ayudantes Culinarios + Bolsa de Hornear', 'cod': 'CABB'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Huevos de Páscua + Chocolates', 'cod': 'MXEC'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Platos Listos Congelados + Pasta', 'cod': 'MXDP'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Platos Congelados y Listos para Comer', 'cod': 'MXFR'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Alimentos Congelados + Margarina', 'cod': 'MXFM'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Modificadores + Cocoa', 'cod': 'MXMC'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Pastas', 'cod': 'MXPS'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Sopas+Cremas+Ramen', 'cod': 'MXSO'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Margarina + Mayonesa + Queso Crema', 'cod': 'MXSP'},
    {'categoria': 'Alimentos', 'descripcion': 'Mixta Azúcar+Endulzantes', 'cod': 'MXSW'},
    {'categoria': 'Alimentos', 'descripcion': 'Mostaza', 'cod': 'MUST'},
    {'categoria': 'Alimentos', 'descripcion': 'Sustitutos de Crema', 'cod': 'NDCR'},
    {'categoria': 'Alimentos', 'descripcion': 'Fideos', 'cod': 'NOOD'},
    {'categoria': 'Alimentos', 'descripcion': 'Nuggets', 'cod': 'NUGG'},
    {'categoria': 'Alimentos', 'descripcion': 'Avena en hojuelas_liquidas', 'cod': 'OAFL'},
    {'categoria': 'Alimentos', 'descripcion': 'Aceitunas', 'cod': 'OLIV'},
    {'categoria': 'Alimentos', 'descripcion': 'Tortilla', 'cod': 'PANC'},
    {'categoria': 'Alimentos', 'descripcion': 'Panetón', 'cod': 'PANE'},
    {'categoria': 'Alimentos', 'descripcion': 'Pastas', 'cod': 'PAST'},
    {'categoria': 'Alimentos', 'descripcion': 'Salsas para Pasta', 'cod': 'PSAU'},
    {'categoria': 'Alimentos', 'descripcion': 'Turrón de maní', 'cod': 'PNOU'},
    {'categoria': 'Alimentos', 'descripcion': 'Carne Porcina', 'cod': 'PORK'},
    {'categoria': 'Alimentos', 'descripcion': 'Postres en Polvo_Postres para Preparar _ Horneables-Gelificables', 'cod': 'PPMX'},
    {'categoria': 'Alimentos', 'descripcion': 'Leche de Soya en Polvo', 'cod': 'PWSM'},
    {'categoria': 'Alimentos', 'descripcion': 'Cereales Precocidos', 'cod': 'PCCE'},
    {'categoria': 'Alimentos', 'descripcion': 'Masas Frescas_Tapas Empanadas y Tarta', 'cod': 'DOUG'},
    {'categoria': 'Alimentos', 'descripcion': 'Pre-Pizzas', 'cod': 'PPIZ'},
    {'categoria': 'Alimentos', 'descripcion': 'Meriendas listas', 'cod': 'REFR'},
    {'categoria': 'Alimentos', 'descripcion': 'Arroz', 'cod': 'RICE'},
    {'categoria': 'Alimentos', 'descripcion': 'Galletas de Arroz', 'cod': 'RBIS'},
    {'categoria': 'Alimentos', 'descripcion': 'Frijoles Procesados', 'cod': 'RTEB'},
    {'categoria': 'Alimentos', 'descripcion': 'Pratos Prontos _ Comidas Listas', 'cod': 'RTEM'},
    {'categoria': 'Alimentos', 'descripcion': 'Aderezos para Ensalada', 'cod': 'SDRE'},
    {'categoria': 'Alimentos', 'descripcion': 'Sal', 'cod': 'SALT'},
    {'categoria': 'Alimentos', 'descripcion': 'Galletas Saladas_Galletas No Dulce', 'cod': 'SLTC'},
    {'categoria': 'Alimentos', 'descripcion': 'Sardina Envasada', 'cod': 'SARD'},
    {'categoria': 'Alimentos', 'descripcion': 'Cecinas', 'cod': 'SAUS'},
    {'categoria': 'Alimentos', 'descripcion': 'Milanesas', 'cod': 'SCHN'},
    {'categoria': 'Alimentos', 'descripcion': 'Snacks', 'cod': 'SNAC'},
    {'categoria': 'Alimentos', 'descripcion': 'Fideos Sopa', 'cod': 'SNOO'},
    {'categoria': 'Alimentos', 'descripcion': 'Sopas_Sopas Cremas', 'cod': 'SOUP'},
    {'categoria': 'Alimentos', 'descripcion': 'Siyau', 'cod': 'SOYS'},
    {'categoria': 'Alimentos', 'descripcion': 'Tallarines_Spaguetti', 'cod': 'SPAG'},
    {'categoria': 'Alimentos', 'descripcion': 'Chocolate para Untar', 'cod': 'SPCH'},
    {'categoria': 'Alimentos', 'descripcion': 'Azucar', 'cod': 'SUGA'},
    {'categoria': 'Alimentos', 'descripcion': 'Galletas Dulces', 'cod': 'SWCO'},
    {'categoria': 'Alimentos', 'descripcion': 'Untables Dulces', 'cod': 'SWSP'},
    {'categoria': 'Alimentos', 'descripcion': 'Endulzantes', 'cod': 'SWEE'},
    {'categoria': 'Alimentos', 'descripcion': 'Torradas _ Tostadas', 'cod': 'TOAS'},
    {'categoria': 'Alimentos', 'descripcion': 'Salsas de Tomate', 'cod': 'TOMA'},
    {'categoria': 'Alimentos', 'descripcion': 'Atún Envasado', 'cod': 'TUNA'},
    {'categoria': 'Alimentos', 'descripcion': 'Leche Vegetal', 'cod': 'VMLK'},
    {'categoria': 'Alimentos', 'descripcion': 'Harinas de trigo', 'cod': 'WFLO'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Ambientadores_Desodorante Ambiental', 'cod': 'AIRC'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Jabón en Barra_Jabón de lavar', 'cod': 'BARS'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Cloro_Lavandinas_Lejías_Blanqueadores', 'cod': 'BLEA'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Pastillas para Inodoro', 'cod': 'CBLK'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Guantes de látex', 'cod': 'CGLO'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Esponjas de Limpieza_Esponjas y paños', 'cod': 'CLSP'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Utensilios de Limpieza', 'cod': 'CLTO'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Filtros de Café', 'cod': 'FILT'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Cross Category (Limpiadores Domesticos)', 'cod': 'CRHC'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Cross Category (Lavandería)', 'cod': 'CRLA'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Cross Category (Productos de Papel)', 'cod': 'CRPA'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Lavavajillas_Lavaplatos _ Lavalozas mano', 'cod': 'DISH'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Empaques domésticos_Bolsas plásticas_Plástico Adherente_Papel encerado_Papel aluminio', 'cod': 'DPAC'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Destapacañerias', 'cod': 'DRUB'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Perfumantes para Ropa_Perfumes para Ropa', 'cod': 'FBRF'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Cera p_pisos', 'cod': 'FWAX'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Desodorante para Pies', 'cod': 'FDEO'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Lustramuebles', 'cod': 'FRNP'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Bolsas de Basura', 'cod': 'GBBG'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Limpiadores verdes', 'cod': 'GCLE'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Limpiadores_Limpiadores y Desinfectantes', 'cod': 'CLEA'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Insecticidas_Raticidas', 'cod': 'INSE'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Toallas de papel_Papel Toalla _ Toallas de Cocina _ Rollos Absorbentes de Papel', 'cod': 'KITT'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Detergentes para ropa', 'cod': 'LAUN'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Apresto', 'cod': 'LSTA'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Mixta Pastillas para Inodoro + Limpiadores', 'cod': 'MXBC'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Mixta Home Care_Cloro-Limpiadores-Ceras-Ambientadores', 'cod': 'MXHC'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Mixta Limpiadores + Cloro', 'cod': 'MXCB'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Mixta Detergentes + Cloro', 'cod': 'MXLB'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Mixta Detergentes + Lavavajillas', 'cod': 'MXLD'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Pañitos + Papel Higienico', 'cod': 'CRTO'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Servilletas', 'cod': 'NAPK'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Film plastico e papel aluminio', 'cod': 'PLWF'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Esponjas de Acero', 'cod': 'SCOU'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Suavizantes de Ropa', 'cod': 'SOFT'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Quitamanchas_Desmanchadores', 'cod': 'STRM'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Papel Higiénico', 'cod': 'TOIP'},
    {'categoria': 'Cuidado del Hogar', 'descripcion': 'Paños de Limpieza', 'cod': 'WIPE'},
    {'categoria': 'OTC', 'descripcion': 'Analgésicos_Painkillers', 'cod': 'ANLG'},
    {'categoria': 'OTC', 'descripcion': 'Suplementos alimentares', 'cod': 'FSUP'},
    {'categoria': 'OTC', 'descripcion': 'Gastrointestinales_Efervescentes', 'cod': 'GMED'},
    {'categoria': 'OTC', 'descripcion': 'Vitaminas y Calcio', 'cod': 'VITA'},
    {'categoria': 'Otros', 'descripcion': 'Categoría Desconocida', 'cod': 'nan'},
    {'categoria': 'Otros', 'descripcion': 'Pilas_Baterías', 'cod': 'BATT'},
    {'categoria': 'Otros', 'descripcion': 'Combustible Gas', 'cod': 'CGAS'},
    {'categoria': 'Otros', 'descripcion': 'Panel Financiero de Hogares', 'cod': 'PFIN'},
    {'categoria': 'Otros', 'descripcion': 'Panel Financiero de Hogares', 'cod': 'PFIN'},
    {'categoria': 'Otros', 'descripcion': 'Cartuchos de Tintas', 'cod': 'INKC'},
    {'categoria': 'Otros', 'descripcion': 'Alimento para Mascota_Alim.p _ perro _ gato', 'cod': 'PETF'},
    {'categoria': 'Otros', 'descripcion': 'Telecomunicaciones _ Convergencia', 'cod': 'TELE'},
    {'categoria': 'Otros', 'descripcion': 'Tickets _ Till Rolls', 'cod': 'TILL'},
    {'categoria': 'Otros', 'descripcion': 'Tabaco _ Cigarrillos', 'cod': 'TOBA'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Incontinencia de Adultos', 'cod': 'ADIP'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Shampoo Infantil', 'cod': 'BSHM'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Maquinas de Afeitar', 'cod': 'RAZO'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cremas Corporales', 'cod': 'BDCR'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Paños Húmedos', 'cod': 'CWIP'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cremas para Peinar', 'cod': 'COMB'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Acondicionador_Bálsamo', 'cod': 'COND'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cross Category (Higiene)', 'cod': 'CRHY'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cross Category (Personal Care)', 'cod': 'CRPC'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Desodorantes', 'cod': 'DEOD'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Pañales_Pañales Desechables', 'cod': 'DIAP'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cremas Faciales', 'cod': 'FCCR'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Pañuelos Faciales', 'cod': 'FTIS'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Protección Femenina_Toallas Femeninas', 'cod': 'FEMI'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Fragancias', 'cod': 'FRAG'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cuidado del Cabello_Hair Care', 'cod': 'HAIR'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Tintes para el Cabello_Tintes _ Tintura _ Tintes y Coloración para el cabello', 'cod': 'HRCO'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Depilación', 'cod': 'HREM'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Alisadores para el Cabello', 'cod': 'HRST'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Fijadores para el Cabello_Modeladores_Gel_Fijadores para el cabello', 'cod': 'HSTY'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Tratamientos para el Cabello', 'cod': 'HRTR'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Óleo Calcáreo', 'cod': 'LINI'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Maquillaje_Cosméticos', 'cod': 'MAKE'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Jabón Medicinal', 'cod': 'MEDS'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Mixta Make Up+Tinturas', 'cod': 'MXMH'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Enjuague Bucal_Refrescante Bucal', 'cod': 'MOWA'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cuidado Bucal', 'cod': 'ORAL'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Protectores Femeninos', 'cod': 'SPAD'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Toallas Femininas', 'cod': 'STOW'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Shampoo', 'cod': 'SHAM'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Afeitado_Crema afeitar_Loción de afeitar_Pord. Antes del afeitado', 'cod': 'SHAV'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cremas Faciales y Corporales_Cremas de Belleza _ Cremas Cuerp y Faciales', 'cod': 'SKCR'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Protección Solar', 'cod': 'SUNP'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Talcos_Talco para pies', 'cod': 'TALC'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Tampones Femeninos', 'cod': 'TAMP'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Jabón de Tocador', 'cod': 'TOIL'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Cepillos Dentales', 'cod': 'TOOB'},
    {'categoria': 'Cuidado Personal', 'descripcion': 'Pastas Dentales', 'cod': 'TOOT'},
    {'categoria': 'Material Escolar', 'descripcion': 'Morrales y MAletas Escoalres', 'cod': 'BAGS'},
    {'categoria': 'Material Escolar', 'descripcion': 'Lapices de Colores', 'cod': 'CLPC'},
    {'categoria': 'Material Escolar', 'descripcion': 'Lapices De Grafito', 'cod': 'GRPC'},
    {'categoria': 'Material Escolar', 'descripcion': 'Unidaddores', 'cod': 'MRKR'},
    {'categoria': 'Material Escolar', 'descripcion': 'Cuadernos', 'cod': 'NTBK'},
    {'categoria': 'Material Escolar', 'descripcion': 'Útiles Escolares', 'cod': 'SCHS'},
    {'categoria': 'Diversos', 'descripcion': 'Estudio de Categorías', 'cod': 'CSTD'},
    {'categoria': 'Diversos', 'descripcion': 'Corporativa', 'cod': 'CORP'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category', 'cod': 'CROS'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Bebés)', 'cod': 'CRBA'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Desayuno)_Yogurt, Cereal, Pan y Queso', 'cod': 'CRBR'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Diet y Light)', 'cod': 'CRDT'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Alimentos Secos)', 'cod': 'CRDF'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Alimentos)', 'cod': 'CRFO'},
    {'categoria': 'Diversos', 'descripcion': 'Cross Category (Salsas)_Mayonesas-Ketchup _ Salsas Frías', 'cod': 'CRSA'},
    {'categoria': 'Diversos', 'descripcion': 'Demo', 'cod': 'DEMO'},
    {'categoria': 'Diversos', 'descripcion': 'Flash', 'cod': 'FLSH'},
    {'categoria': 'Diversos', 'descripcion': 'Holistic View', 'cod': 'HLVW'},
    {'categoria': 'Diversos', 'descripcion': 'Mezcla para café instantaneo y crema no láctea', 'cod': 'COCP'},
    {'categoria': 'Diversos', 'descripcion': 'Mezclas nutricionales y suplementos', 'cod': 'CRSN'},
    {'categoria': 'Diversos', 'descripcion': 'Consolidado_Multicategory', 'cod': 'MULT'},
    {'categoria': 'Diversos', 'descripcion': 'Pantry Check', 'cod': 'PCHK'},
    {'categoria': 'Diversos', 'descripcion': 'Inventario', 'cod': 'STCK'},
    {'categoria': 'Diversos', 'descripcion': 'Leche y Cereales Calientes_Cereales Precocidos y Leche Líquida Blanca', 'cod': 'MIHC'}
]

TEMPLATE_FILE_NAME = "Plantilla_Entrada_5W1H.xlsx"
README_SHEET_NAME = "README"
TEMPLATE_SEGMENTS = [
    ("1W - Grafico de tendencia mensual (MAT)", "1_MarcaEjemplo"),
    ("2W - Arbol de medidas", "2_MarcaEjemplo_K"),
    ("3W - Que? Tamaños", "3_MarcaEjemplo_1"),
    ("3W - Que? Marcas", "3_MarcaEjemplo_2"),
    ("3W - Que? Sabores", "3_MarcaEjemplo_3"),
    ("4W - Quienes? NSE", "4_MarcaEjemplo"),
    ("5W - Donde? Regiones", "5_MarcaEjemplo_1"),
    ("5W - Donde? Canales", "5_MarcaEjemplo_2"),
    ("6W - Players", "6_Categoria_XX"),
    ("6W - Precio indexado", "6_Categoria_1"),
    ("7W - Distribucion", "7_Categoria_Canal"),
    ("8W - Intervalos de confianza", "8_Categoria"),
]
QUESTION_STYLE_MAP = {
    "1": "bold #1f77b4",
    "2": "bold #d62728",
    "3": "bold #2ca02c",
    "4": "bold #ff7f0e",
    "5": "bold #9467bd",
    "6": "bold #17becf",
    "7": "bold #8c564b",
    "8": "bold #e377c2",
}
DEFAULT_SEGMENT_STYLE = "bold white"

def question_key_from_template_sheet(template_sheet):
    """Extrae el numero de pregunta base desde el nombre de hoja plantilla."""
    if not isinstance(template_sheet, str):
        return ""
    first_token = template_sheet.split('_', 1)[0].strip()
    return first_token[0] if first_token else ""

def question_style_for_sheet(template_sheet):
    """Devuelve el estilo rich asociado a la pregunta."""
    qkey = question_key_from_template_sheet(template_sheet)
    return QUESTION_STYLE_MAP.get(qkey, DEFAULT_SEGMENT_STYLE)

def buscar_categorias(keyword):
    keyword_normalizado = strip_accents(keyword)
    # Filtra la lista buscando en 'descripcion' y en 'cod'
    return [cat for cat in categorias 
            if keyword_normalizado in strip_accents(cat['descripcion']) 
            or keyword_normalizado in strip_accents(cat['cod'])]

def seleccionar_categoria():
    while True:
        keyword = input(f"{PromptColors.CATEGORY_SEARCH}Ingrese palabra clave para la categoría: {Colors.ENDC}").strip()
        if not keyword:
            print(f"{Colors.FAIL}La palabra clave no puede estar vacía.{Colors.ENDC}")
            continue
        matches = buscar_categorias(keyword)
        if not matches:
            print(f"{Colors.FAIL}No se encontraron categorías. Intente de nuevo.{Colors.ENDC}")
            continue
        
        print(f"{Colors.OKCYAN}Categorías encontradas:{Colors.ENDC}")
        for idx, cat in enumerate(matches, 1):
            print(f"{Colors.OKBLUE}{idx}. {cat['descripcion']} ({cat['cod']}){Colors.ENDC}")
        seleccion = input(f"{PromptColors.CATEGORY_SELECT}Número de categoría (o 'r' para reintentar): {Colors.ENDC}").strip()
        if seleccion.lower() == 'r':
            continue
        
        try:
            seleccion = int(seleccion)
            if 1 <= seleccion <= len(matches):
                return matches[seleccion - 1]
            else:
                print(f"{Colors.FAIL}Selección inválida.{Colors.ENDC}")
        except ValueError:
            print(f"{Colors.FAIL}Entrada inválida. Use números o 'r'.{Colors.ENDC}")

def sanitizar_nombre_hoja(nombre, max_len=31):
    """Sanitiza nombre de hoja para Excel (caracteres invalidos + longitud)."""
    texto = str(nombre).strip()
    for char in ['\\', '/', '*', '[', ']', ':', '?']:
        texto = texto.replace(char, '_')
    texto = ' '.join(texto.split()).strip(" '")
    if not texto:
        texto = "Hoja"
    return texto[:max_len]

def sanitizar_segmento_archivo(texto, max_len=80):
    """Sanitiza texto para uso en nombre de archivo."""
    value = str(texto).strip()
    invalid_chars = '<>:"/\\|?*'
    value = ''.join('_' if (ch in invalid_chars or ord(ch) < 32) else ch for ch in value)
    value = ' '.join(value.split()).strip(" .")
    if not value:
        value = "NA"
    return value[:max_len]

def generar_nombre_unico(nombre_archivo):
    if not os.path.exists(nombre_archivo):
        return nombre_archivo
    base, ext = os.path.splitext(nombre_archivo)
    contador = 1
    while True:
        nuevo_nombre = f"{base} ({contador}){ext}"
        if not os.path.exists(nuevo_nombre):
            return nuevo_nombre
        contador += 1

def obtener_marcas():
    """Solicita una lista de marcas (una por linea, Enter vacio para terminar)."""
    marcas = []
    vistos = set()
    print(f"{Colors.OKCYAN}Ingrese las marcas (Enter vacio para terminar):{Colors.ENDC}")
    while True:
        marca = input(f"{PromptColors.BRAND}Marca #{len(marcas) + 1}: {Colors.ENDC}").strip()
        if not marca:
            if marcas:
                return marcas
            print(f"{Colors.FAIL}Debe ingresar al menos una marca.{Colors.ENDC}")
            continue
        norm = strip_accents(marca)
        if norm in vistos:
            print(f"{Colors.WARNING}La marca '{marca}' ya fue ingresada; se omitira duplicado.{Colors.ENDC}")
            continue
        vistos.add(norm)
        marcas.append(marca)

def mostrar_opciones_plantillas():
    print(f"{Colors.OKCYAN}Plantillas disponibles:{Colors.ENDC}")
    print("1. TODAS", style="bold #ECEFF1", markup=False)
    for idx, (label, template_sheet) in enumerate(TEMPLATE_SEGMENTS, start=2):
        style = question_style_for_sheet(template_sheet)
        print(f"{idx}. {label} ({template_sheet})", style=style, markup=False)

def mostrar_segmentos_seleccionados(hojas_seleccionadas):
    """Muestra el resumen de segmentos elegidos respetando color por pregunta."""
    print(f"{Colors.OKCYAN}Segmentos seleccionados:{Colors.ENDC}")
    selected_set = set(hojas_seleccionadas)
    shown = 0
    for label, sheet in TEMPLATE_SEGMENTS:
        if sheet in selected_set:
            style = question_style_for_sheet(sheet)
            print(f"  - {label} ({sheet})", style=style, markup=False)
            shown += 1
    if shown == 0:
        print(f"{Colors.WARNING}No hay segmentos seleccionados.{Colors.ENDC}")

def seleccionar_plantillas():
    """Devuelve lista de hojas plantilla a conservar segun seleccion del usuario."""
    option_to_sheet = {idx: sheet for idx, (_, sheet) in enumerate(TEMPLATE_SEGMENTS, start=2)}
    while True:
        mostrar_opciones_plantillas()
        raw = input(
            f"{PromptColors.TEMPLATE_SELECT}Opciones (1=TODAS | ej: 2,4,9 | Enter=TODAS): {Colors.ENDC}"
        ).strip()
        if not raw:
            return [sheet for _, sheet in TEMPLATE_SEGMENTS]
        tokens = [tok for tok in re.split(r"[,\s]+", raw) if tok]
        try:
            selected_numbers = [int(tok) for tok in tokens]
        except ValueError:
            print(f"{Colors.FAIL}Entrada invalida. Use solo numeros separados por coma.{Colors.ENDC}")
            continue
        valid_numbers = set(option_to_sheet.keys()) | {1}
        if any(num not in valid_numbers for num in selected_numbers):
            print(f"{Colors.FAIL}Hay opciones fuera de rango. Intente nuevamente.{Colors.ENDC}")
            continue
        if 1 in selected_numbers:
            return [sheet for _, sheet in TEMPLATE_SEGMENTS]
        selected_sheets = []
        for num in selected_numbers:
            sheet = option_to_sheet[num]
            if sheet not in selected_sheets:
                selected_sheets.append(sheet)
        if selected_sheets:
            return selected_sheets
        print(f"{Colors.FAIL}Debe seleccionar al menos una plantilla.{Colors.ENDC}")

def solicitar_etiqueta_categoria(cat_sel):
    default_value = str(cat_sel.get('descripcion', 'Categoria')).strip() or 'Categoria'
    raw = input(
        f"{PromptColors.CATEGORY_LABEL}Etiqueta para reemplazar 'Categoria' en hojas (Enter para '{default_value}'): {Colors.ENDC}"
    ).strip()
    return raw if raw else default_value

def solicitar_etiqueta_players():
    raw = input(
        f"{PromptColors.PLAYERS_LABEL}Etiqueta objetivo para Players (hoja 6_Categoria_XX, ej: Fabricante/Marca Propia). Enter para 'XX': {Colors.ENDC}"
    ).strip()
    return raw if raw else 'XX'

def solicitar_corte_distribucion():
    raw = input(
        f"{PromptColors.DISTRIBUTION_LABEL}Etiqueta para corte de distribucion (hoja 7_...; Enter para 'Canal'): {Colors.ENDC}"
    ).strip()
    return raw if raw else 'Canal'

def solicitar_unidad_arbol():
    """
    Solicita la unidad de medida para Segmento 2 (arbol de medidas).
    Devuelve (letra, nombre_unidad).
    """
    print(f"{Colors.OKCYAN}Unidad para Segmento 2 (Arbol de medidas):{Colors.ENDC}")
    unit_options = list(TREE_UNIT_MAP.items())
    index_lookup = {str(idx): key for idx, (key, _) in enumerate(unit_options, start=1)}
    for idx, (key, label) in enumerate(unit_options, start=1):
        print(f"{Colors.OKBLUE}  {idx}= {key} -> {label}{Colors.ENDC}")
    while True:
        raw = input(
            f"{PromptColors.TREE_UNIT}Unidad para hoja 2_* (numero o letra). Enter para 'K' (Kilos): {Colors.ENDC}"
        ).strip().upper()
        if not raw:
            return "K", TREE_UNIT_MAP["K"]
        if raw in index_lookup:
            unit_key = index_lookup[raw]
            return unit_key, TREE_UNIT_MAP[unit_key]
        if raw in TREE_UNIT_MAP:
            return raw, TREE_UNIT_MAP[raw]
        print(f"{Colors.FAIL}Unidad invalida. Use numero (1-{len(unit_options)}) o letra U/L/K/T/R/M/H.{Colors.ENDC}")

def solicitar_fabricante():
    """Solicita fabricante para el nombre del archivo de salida."""
    while True:
        fabricante = input(f"{PromptColors.MANUFACTURER}Ingrese nombre del fabricante (para el nombre del archivo): {Colors.ENDC}").strip()
        if fabricante:
            return fabricante
        print(f"{Colors.FAIL}El fabricante no puede estar vacio.{Colors.ENDC}")

def asegurar_nombre_hoja_unico(nombre, usados):
    """Garantiza nombre unico de hoja respetando maximo 31 caracteres."""
    base = sanitizar_nombre_hoja(nombre)
    if base not in usados:
        return base
    contador = 1
    while True:
        sufijo = f"_{contador}"
        disponible = 31 - len(sufijo)
        candidato = f"{base[:disponible]}{sufijo}"
        if candidato not in usados:
            return candidato
        contador += 1

def construir_nombre_hoja(template_name, marca, categoria_label, players_suffix, distribution_cut, tree_unit_letter="K"):
    """Construye el nombre final de hoja desde el nombre plantilla."""
    title = template_name.replace('MarcaEjemplo', marca).replace('Categoria', categoria_label)
    if template_name.startswith('2_'):
        title = re.sub(r'_[A-Za-z]$', f"_{tree_unit_letter}", title)
    if template_name.endswith('_XX'):
        title = title[:-2] + players_suffix
    if template_name.startswith('7_'):
        parts = title.split('_')
        if len(parts) >= 3:
            parts[-1] = distribution_cut
            title = '_'.join(parts)
    return sanitizar_nombre_hoja(title)

def aplicar_reemplazos_en_celdas(sheet, replacements):
    """Reemplaza placeholders textuales en celdas string."""
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                updated = cell.value
                for old, new in replacements.items():
                    updated = updated.replace(old, new)
                if updated != cell.value:
                    cell.value = updated

def agregar_resumen_parametros(wb, summary_lines):
    """
    Agrega al final del archivo un resumen de parametros seleccionados.
    Se escribe al final de la hoja README si existe.
    """
    if not summary_lines:
        return
    if README_SHEET_NAME in wb.sheetnames:
        ws = wb[README_SHEET_NAME]
    else:
        ws = wb.create_sheet("PARAMETROS_GENERACION")
    last_row = ws.max_row if ws.max_row else 1
    start_row = last_row + 2
    ws.cell(row=start_row, column=1, value="Parametros seleccionados")
    for idx, line in enumerate(summary_lines, start=1):
        ws.cell(row=start_row + idx, column=1, value=line)

def crear_excel_desde_plantilla(
    nombre_archivo,
    marcas,
    categoria_label,
    nombre_pais,
    hojas_seleccionadas,
    players_suffix,
    distribution_cut,
    tree_unit_letter="K",
    summary_lines=None,
):
    """Crea un archivo Excel desde Plantilla_Entrada_5W1H con contenido filtrado y dinamico."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("Falta openpyxl. Instale con: pip install openpyxl") from exc

    template_path = os.path.join(os.getcwd(), TEMPLATE_FILE_NAME)
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"No se encontro el archivo plantilla '{TEMPLATE_FILE_NAME}' en {os.getcwd()}")

    nombre_archivo_unico = generar_nombre_unico(nombre_archivo)
    wb = load_workbook(template_path)
    existing_sheets = list(wb.sheetnames)

    keep_sheets = []
    if README_SHEET_NAME in existing_sheets:
        keep_sheets.append(README_SHEET_NAME)
    selected_found = 0
    for sheet in hojas_seleccionadas:
        if sheet in existing_sheets and sheet not in keep_sheets:
            keep_sheets.append(sheet)
            selected_found += 1
        elif sheet not in existing_sheets:
            print(f"{Colors.WARNING}Aviso: la hoja plantilla '{sheet}' no existe y se omitira.{Colors.ENDC}")

    if selected_found == 0:
        raise ValueError("No se encontraron hojas de trabajo seleccionadas dentro de la plantilla.")

    for title in list(wb.sheetnames):
        if title not in keep_sheets:
            wb.remove(wb[title])

    usados = set()
    marcas_validas = [m for m in marcas if str(m).strip()]
    if not marcas_validas:
        raise ValueError("Debe proporcionar al menos una marca.")
    ordered_templates = [
        sheet
        for _, sheet in TEMPLATE_SEGMENTS
        if sheet in keep_sheets and sheet != README_SHEET_NAME
    ]
    brand_template_names = [name for name in ordered_templates if 'MarcaEjemplo' in name]
    category_template_names = [name for name in ordered_templates if 'MarcaEjemplo' not in name]

    generated_brand_sheets = {}
    generated_category_sheets = {}

    # Genera hojas de marca para todas las marcas sin intercalar contenido.
    for template_name in brand_template_names:
        ws = wb[template_name]
        sheets_for_brand = [ws]
        for _ in marcas_validas[1:]:
            sheets_for_brand.append(wb.copy_worksheet(ws))
        generated_brand_sheets[template_name] = []
        for idx, brand_sheet in enumerate(sheets_for_brand):
            marca_actual = marcas_validas[idx]
            replacements = {
                'MarcaEjemplo': marca_actual,
                'CategoriaEjemplo': categoria_label,
                'PaisEjemplo': nombre_pais,
            }
            final_name = construir_nombre_hoja(
                template_name,
                marca_actual,
                categoria_label,
                players_suffix,
                distribution_cut,
                tree_unit_letter
            )
            final_name = asegurar_nombre_hoja_unico(final_name, usados)
            brand_sheet.title = final_name
            usados.add(brand_sheet.title)
            aplicar_reemplazos_en_celdas(brand_sheet, replacements)
            generated_brand_sheets[template_name].append(brand_sheet)

    # Hojas no-marca (categoria/players/distribucion/intervalos).
    for template_name in category_template_names:
        ws = wb[template_name]
        replacements = {
            'MarcaEjemplo': marcas_validas[0],
            'CategoriaEjemplo': categoria_label,
            'PaisEjemplo': nombre_pais,
        }
        final_name = construir_nombre_hoja(
            template_name,
            marcas_validas[0],
            categoria_label,
            players_suffix,
            distribution_cut,
            tree_unit_letter
        )
        final_name = asegurar_nombre_hoja_unico(final_name, usados)
        ws.title = final_name
        usados.add(ws.title)
        aplicar_reemplazos_en_celdas(ws, replacements)
        generated_category_sheets[template_name] = ws

    # Orden final: README -> 6/7/8 -> (por marca) 1,2,3,4,5...
    desired_sheets = []
    if README_SHEET_NAME in wb.sheetnames:
        desired_sheets.append(wb[README_SHEET_NAME])

    def _category_priority(template_name: str) -> tuple[int, int]:
        order_idx = ordered_templates.index(template_name) if template_name in ordered_templates else 999
        if template_name.startswith('6_'):
            return (0, order_idx)
        if template_name.startswith('7_'):
            return (1, order_idx)
        if template_name.startswith('8_'):
            return (2, order_idx)
        return (3, order_idx)

    for template_name in sorted(category_template_names, key=_category_priority):
        sheet_obj = generated_category_sheets.get(template_name)
        if sheet_obj is not None:
            desired_sheets.append(sheet_obj)

    for brand_idx, _ in enumerate(marcas_validas):
        for template_name in brand_template_names:
            sheet_list = generated_brand_sheets.get(template_name, [])
            if brand_idx < len(sheet_list):
                desired_sheets.append(sheet_list[brand_idx])

    # Conserva cualquier hoja remanente no contemplada al final (fallback seguro).
    desired_ids = {id(ws) for ws in desired_sheets}
    for ws in wb.worksheets:
        if id(ws) not in desired_ids:
            desired_sheets.append(ws)
            desired_ids.add(id(ws))
    wb._sheets = desired_sheets

    agregar_resumen_parametros(wb, summary_lines)
    wb.save(nombre_archivo_unico)

    if nombre_archivo_unico == nombre_archivo:
        msg = f"Archivo de Excel '{nombre_archivo_unico}' creado exitosamente."
    else:
        msg = (f"El archivo '{nombre_archivo}' ya existia.\n"
               f"Se creo un nuevo archivo '{nombre_archivo_unico}'.")
    print(f"{Colors.OKGREEN}{msg}{Colors.ENDC}")
    return nombre_archivo_unico

def limpiar_pantalla():
    # No limpiamos la terminal para conservar el historial completo.
    return

def mostrar_encabezado(contador, lista_archivos):
    console.rule("[bold bright_cyan]Automatizador de Archivos[/bold bright_cyan]")
    print(f"{Colors.OKCYAN}Archivos creados: {contador}{Colors.ENDC}")
    if lista_archivos:
        print(f"{Colors.OKCYAN}Lista de archivos creados:{Colors.ENDC}")
        for idx, archivo in enumerate(lista_archivos, 1):
            print(f"  {idx}. {archivo}")
    print("\n" + "-"*50 + "\n")

def main():
    contador = 0
    lista_archivos = []

    try:
        while True:
            limpiar_pantalla()
            mostrar_encabezado(contador, lista_archivos)
            print(f"{Colors.OKGREEN}Nota: Puede salir del programa en cualquier momento presionando Ctrl+C.{Colors.ENDC}\n")
            
            # 1. Seleccionar país
            while True:
                input_pais = input(f"{PromptColors.COUNTRY}Ingrese nombre/abreviación del país: {Colors.ENDC}").strip()
                if not input_pais:
                    print(f"{Colors.FAIL}No puede estar vacío.{Colors.ENDC}")
                    continue
                codigo_pais, nombre_pais = obtener_codigo_pais(input_pais)
                if codigo_pais:
                    print(f"{Colors.OKCYAN}País: {nombre_pais} (Código: {codigo_pais}){Colors.ENDC}")
                    cobertura = pop_coverage.get(nombre_pais)
                    if cobertura:
                        print(f"{Colors.OKBLUE}Cobertura poblacional: {cobertura}{Colors.ENDC}")
                    print()
                    break
                else:
                    print(f"{Colors.FAIL}País no encontrado. Intente nuevamente.{Colors.ENDC}\n")
            
            # 2. Seleccionar categoría
            cat_sel = seleccionar_categoria()
            print(f"{Colors.OKCYAN}Categoría: {cat_sel['descripcion']} (Código: {cat_sel['cod']}){Colors.ENDC}\n")

            # 3. Solicitar fabricante para nombre del archivo final
            fabricante = solicitar_fabricante()
            print(f"{Colors.OKCYAN}Fabricante para nombre de archivo: {fabricante}{Colors.ENDC}\n")
            
            # 4. Capturar marcas para generar una plantilla unica multimarca
            marcas = obtener_marcas()
            print(f"{Colors.OKCYAN}Se agregaran {len(marcas)} marca(s) en un solo archivo: {', '.join(marcas)}{Colors.ENDC}\n")

            # 5. Seleccionar segmentos/plantillas
            hojas_seleccionadas = seleccionar_plantillas()
            print(f"{Colors.OKBLUE}Plantillas seleccionadas: {len(hojas_seleccionadas)}{Colors.ENDC}\n")

            # 6. Unidad para Segmento 2 (inmediatamente despues de elegir plantillas)
            tree_unit_letter = 'K'
            tree_unit_name = TREE_UNIT_MAP.get(tree_unit_letter, 'Kilos')
            if "2_MarcaEjemplo_K" in hojas_seleccionadas:
                tree_unit_letter, tree_unit_name = solicitar_unidad_arbol()

            # 7. Parametros dinamicos para placeholders de categoria y cortes
            categoria_default = str(cat_sel.get('descripcion', 'Categoria')).strip() or 'Categoria'
            categoria_label = categoria_default
            if any('Categoria' in sheet for sheet in hojas_seleccionadas):
                categoria_label = solicitar_etiqueta_categoria(cat_sel)

            players_suffix = 'XX'
            if "6_Categoria_XX" in hojas_seleccionadas:
                players_suffix = solicitar_etiqueta_players()

            distribution_cut = 'Canal'
            if "7_Categoria_Canal" in hojas_seleccionadas:
                distribution_cut = solicitar_corte_distribucion()

            summary_lines = [
                f"Etiqueta para reemplazar 'Categoria' en hojas (Enter para '{categoria_default}'): {categoria_label}",
                f"Etiqueta objetivo para Players (hoja 6_Categoria_XX, ej: Fabricante/Marca Propia). Enter para 'XX': {players_suffix}",
                f"Etiqueta para corte de distribucion (hoja 7_...; Enter para 'Canal'): {distribution_cut}",
                f"Unidad Segmento 2 (hoja 2_*): {tree_unit_letter} -> {tree_unit_name}",
            ]

            # 7. Generar un unico archivo con todas las marcas
            segmento_nombre = sanitizar_segmento_archivo(fabricante, 80)
            nombre_archivo = f"{codigo_pais}_{cat_sel['cod']}_{segmento_nombre}.xlsx"
            print(f"{Colors.OKBLUE}Generando archivo: {nombre_archivo}{Colors.ENDC}")
            nombre_archivo_creado = crear_excel_desde_plantilla(
                nombre_archivo=nombre_archivo,
                marcas=marcas,
                categoria_label=categoria_label,
                nombre_pais=nombre_pais,
                hojas_seleccionadas=hojas_seleccionadas,
                players_suffix=players_suffix,
                distribution_cut=distribution_cut,
                tree_unit_letter=tree_unit_letter,
                summary_lines=summary_lines,
            )
            contador += 1
            lista_archivos.append(nombre_archivo_creado)
            
            # Esperar a que el usuario esté listo para continuar
            input(f"\n{PromptColors.CONTINUE}Presione Enter para crear otro archivo o Ctrl+C para salir...{Colors.ENDC}\n")
    
    except KeyboardInterrupt:
        limpiar_pantalla()
        mostrar_encabezado(contador, lista_archivos)
        print(f"{Colors.OKCYAN}Programa finalizado por el usuario. ¡Hasta luego!{Colors.ENDC}\n")
        sys.exit()

if __name__ == "__main__":
    main()
